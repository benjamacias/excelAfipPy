import os
import re
import logging
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from os import scandir, getcwd

# Configuración de logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Ruta de archivos recibidos y terminados
DIR_ENTRADA = "recibidos"
DIR_SALIDA = "terminado"
ARCHIVO_CLIENTES = "clientes.xls"

# Diccionarios de mapeo de tipos y columnas
TIPOS_CELDAS = {
    "1 - Factura A": 'S',
    "8 - Nota de Crédito B": 'T',
    "6 - Factura B": 'U',
    "11 - Factura C": 'V',
    "3 - Nota de Crédito A": 'W',
    "201 - Factura de Crédito Electrónica MyPyMEs (FCE) A": 'X',
    "12 - Nota de Débito C": 'Y',
}

COLUMNAS_RESUMEN = {
    'Imp. Neto Gravado': 18,
    'Imp. Neto No Gravado': 19,
    'IVA': 20,
    'Imp. Total': 21,
}


def listar_archivos(ruta=DIR_ENTRADA):
    """Lista archivos en la carpeta de entrada."""
    return [arch.name for arch in scandir(ruta) if arch.is_file()]

def cargar_archivo(path):
    """Carga un archivo Excel o CSV a un DataFrame."""
    if path.lower().endswith(".csv"):
        return pd.read_csv(path)
    else:
        return pd.read_excel(path, header=1)

def sumar_por_tipo(df, tipo, columna):
    """Suma los valores de una columna para un tipo específico de comprobante."""
    return round(df[df["Tipo"] == tipo][columna].fillna(0).sum(), 2)

def obtener_fecha_desde_excel(ws):
    """Extrae y formatea la fecha desde la celda A3."""
    try:
        valor = ws['A3'].value
        if isinstance(valor, datetime):
            return valor.strftime('%m/%Y')
        else:
            return datetime.strptime(str(valor), '%d/%m/%Y').strftime('%m/%Y')
    except Exception:
        return '00/0000'

def obtener_cliente_nombre(cuil_archivo, df_clientes):
    """Obtiene el nombre del cliente a partir del CUIL en el nombre del archivo."""
    try:
        cuil = int(re.search(r'\d{11}', cuil_archivo).group())
        nombre = df_clientes.loc[df_clientes['USUARIOS'] == cuil, 'CLIENTES'].values[0]
        return nombre
    except Exception:
        return None


def procesar_archivos():
    archivos = listar_archivos()
    clientes_excel = pd.read_excel(ARCHIVO_CLIENTES, sheet_name="VERO2023")
    
    for nombre_archivo in archivos:
        ruta = os.path.join(DIR_ENTRADA, nombre_archivo)
        logging.info(f"Procesando archivo: {ruta}")

        try:
            # Cargar archivo de comprobantes
            df = cargar_archivo(ruta)
            df.fillna(0, inplace=True)
            df["Tipo"] = df["Tipo"].astype(str)

            # Cargar workbook de Excel
            wb = load_workbook(ruta)
            ws = wb.active

            # Fecha para el nombre final
            fecha = obtener_fecha_desde_excel(ws)

            # Escribir encabezados en hoja
            encabezados_columnas(ws)

            # Calcular y escribir totales por tipo y columna
            for tipo, col in TIPOS_CELDAS.items():
                for campo, fila in COLUMNAS_RESUMEN.items():
                    total = sumar_por_tipo(df, tipo, campo)
                    ws[f"{col}{fila}"] = f"{total:.2f}"

            # Intentar calcular valor IVA (última fila)
            ws['Z17'] = "Coef. IVA último comprobante"
            ws['Z18'] = calcular_coef_iva_final(df)

            # Guardar archivo con nuevo nombre
            cliente_nombre = obtener_cliente_nombre(nombre_archivo, clientes_excel)
            tipo_libro = "Libro venta " if 'Mis Comprobantes Emitidos' in nombre_archivo else "Libro compra "
            nombre_final = f"{tipo_libro}{cliente_nombre or nombre_archivo} {fecha.replace('/', '-')}.xlsx"
            ruta_final = os.path.join(DIR_SALIDA, nombre_final)

            wb.save(ruta_final)
            logging.info(f"Guardado: {ruta_final}")
        
        except Exception as e:
            logging.error(f"Error procesando {nombre_archivo}: {e}")

def encabezados_columnas(ws):
    """Escribe los encabezados en las celdas fijas del Excel."""
    ws['S17'] = "Factura A"
    ws['T17'] = "Nota de Crédito B"
    ws['U17'] = "Factura B"
    ws['V17'] = "Factura C"
    ws['W17'] = "Nota de Crédito A"
    ws['X17'] = "Factura de Crédito Electrónica MyPyMEs A"
    ws['Y17'] = "12 - Nota de Débito C"
    ws['R18'] = "Imp. Neto Gravado"
    ws['R19'] = "Imp. Neto No Gravado"
    ws['R20'] = "IVA"
    ws['R21'] = "TOTAL"

def calcular_coef_iva_final(df):
    """Calcula el coeficiente IVA del último comprobante válido."""
    for idx in reversed(df.index):
        try:
            imp_total = df.at[idx, "Imp. Total"]
            neto = df.at[idx, "Imp. Neto Gravado"]
            if neto > 0:
                return f"{imp_total / neto:.2f}"
        except Exception:
            continue
    return "0.00"


if __name__ == "__main__":
    procesar_archivos()
 